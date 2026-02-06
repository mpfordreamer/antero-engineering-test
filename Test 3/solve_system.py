import pandas as pd
import numpy as np
from scipy.integrate import solve_ivp
import openpyxl
import os
import sys

# Define Files
INPUT_FILE = "system_inputs.xlsx"
INPUT_SHEET = "Inputs"
OUTPUT_FILE = "system_outputs.xlsx"
OUTPUT_SHEET = "All_Results"

def load_inputs(filename):
    """
    Reads the 'Inputs' sheet as a DataFrame.
    Returns the full DataFrame.
    """
    if not os.path.exists(filename):
        raise FileNotFoundError(f"File {filename} not found.")
    
    df = pd.read_excel(filename, sheet_name=INPUT_SHEET)
    if df.empty:
        raise ValueError("Input sheet is empty.")
    return df

def tff_odes(t, y, Q_f, Q_p, beta, alpha, k_d):
    """
    Defines the system of ODEs.
    """
    C_s, C_d, V_r_current = y
    
    # Safety clamp
    if V_r_current <= 0:
        return [0, 0, 0]
    
    # Calculate Active Volume Vs
    V_s = (1 - alpha) * V_r_current
    
    if V_s <= 1e-6: 
        dCs_dt = 0
    else:
        # C_r calculation logic
        F = Q_f / (Q_f - Q_p)
        
        denom = (1 - F * beta)
        if denom == 0:
            C_r = C_s 
        else:
            C_r = C_s * ( F * (1 - beta) / denom )
            
        term_advection = (Q_f * (1 - beta) / V_s) * (C_r - C_s)
        term_diffusion = (alpha / (1 - alpha)) * k_d * (C_s - C_d)
        
        dCs_dt = term_advection - term_diffusion

    dCd_dt = k_d * (C_s - C_d)
    dVr_dt = -Q_p
    
    return [dCs_dt, dCd_dt, dVr_dt]

def run_simulation():
    print("Loading Inputs...")
    try:
        input_df = load_inputs(INPUT_FILE)
    except Exception as e:
        print(f"Error loading inputs: {e}")
        return

    all_results = []
    
    print(f"Found {len(input_df)} scenarios. Starting Batch Processing...")
    
    # Iterate through rows
    for index, row in input_df.iterrows():
        scenario_id = row["Scenario_ID"]
        print(f"  > Processing: {scenario_id}")
        
        try:
            # Auto-convert comma decimals to dots (for Indonesian Excel)
            def clean_val(val):
                """Converts '0,5' to 0.5 (float)"""
                if isinstance(val, str):
                    val = val.replace(',', '.')
                return float(val)
            # Extract Params
            V_r_0 = clean_val(row["Volume_Retentat_Initial_L"])
            C_s_0 = clean_val(row["Conc_Active_Initial_gL"])
            C_d_0 = clean_val(row["Conc_DeadZone_Initial_gL"])
            Q_f   = clean_val(row["Feed_Flow_Rate_Lmin"])
            Q_p   = clean_val(row["Permeate_Flow_Rate_Lmin"])
            beta  = clean_val(row["Bypass_Fraction_Beta"])
            alpha = clean_val(row["DeadZone_Fraction_Alpha"])
            k_d   = clean_val(row["Mass_Transfer_Coeff_kd"])
            t_end = clean_val(row["Simulation_Time_End_min"])
            dt    = clean_val(row["Time_Step_dt_min"])
            
            # Physics validation: prevent graph explosion if volume runs out
            max_possible_time = V_r_0 / Q_p if Q_p > 0 else 9999
            if t_end > max_possible_time:
                print(f"    [WARNING] Volume runs out at {max_possible_time:.1f} min, limiting simulation")
                t_end = max_possible_time - 1 

            # Solve
            t_span = (0, t_end)
            t_eval = np.arange(0, t_end + dt, dt)
            y0 = [C_s_0, C_d_0, V_r_0]
            
            sol = solve_ivp(
                fun=lambda t, y: tff_odes(t, y, Q_f, Q_p, beta, alpha, k_d),
                t_span=t_span,
                y0=y0,
                t_eval=t_eval,
                method='RK45'
            )
            
            # Post Process
            time = sol.t
            C_s = sol.y[0]
            C_d = sol.y[1]
            V_r = sol.y[2]
            
            # Derived Variables
            F = Q_f / (Q_f - Q_p)
            denom = (1 - F * beta)
            if denom == 0: denom = 1e-9
            C_r_vec = C_s * (F * (1 - beta) / denom)
            C_total = (1 - beta) * C_s + beta * C_r_vec
            
            # Create DataFrame
            df_res = pd.DataFrame({
                "Scenario_ID": scenario_id,
                "Time_min": time,
                "Volume_Retentat_L": V_r,
                "Conc_Active_gL": C_s,
                "Conc_DeadZone_gL": C_d,
                "Conc_Total_Output_gL": C_total
            })
            
            all_results.append(df_res)
            
        except Exception as e:
            print(f"    ERROR in scenario {scenario_id}: {e}")
            
    # Combined Results for Dashboard
    if not all_results:
        print("No results generated.")
        return
        
    final_df = pd.concat(all_results, ignore_index=True)
    
    # OUTPUT FILE WRITING
    print(f"Writing results to {OUTPUT_FILE}...")
    
    try:
        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', mode='w') as writer:
            
            # Write Individual Scenario Sheets
            for df_res in all_results:
                # CAST TO FLOAT: Ensure numeric columns are strictly numbers
                cols_to_convert = ["Time_min", "Volume_Retentat_L", "Conc_Active_gL", "Conc_DeadZone_gL", "Conc_Total_Output_gL"]
                for col in cols_to_convert:
                    df_res[col] = pd.to_numeric(df_res[col], errors='coerce')
                
                # Sheet Name Logic
                sheet_name = str(df_res["Scenario_ID"].iloc[0])
                if len(sheet_name) > 31: sheet_name = sheet_name[:31]
                
                df_res.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"  > Sheet created: {sheet_name}")

            # Write Summary Sheet (hidden, used for charts only)
            final_df = final_df.sort_values(by=["Scenario_ID", "Time_min"])
            for col in ["Time_min", "Volume_Retentat_L", "Conc_Active_gL", "Conc_DeadZone_gL", "Conc_Total_Output_gL"]:
                final_df[col] = pd.to_numeric(final_df[col], errors='coerce')
            final_df.to_excel(writer, sheet_name="_ChartData", index=False)

        # Add Charts (Post-Processing)
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        
        # Hide the chart data sheet
        ws_data = wb["_ChartData"]
        ws_data.sheet_state = 'hidden'
        
        # Create Dashboard Sheet
        if "Dashboard" in wb.sheetnames: del wb["Dashboard"]
        ws_dash = wb.create_sheet("Dashboard", 0) 
        
        # Clean Dashboard
        ws_dash.sheet_view.showGridLines = False
        
        # Add title row
        ws_dash['B1'] = "TFF Simulation Dashboard"
        ws_dash['B1'].font = openpyxl.styles.Font(size=16, bold=True)
        
        # Line Chart
        chart_line = openpyxl.chart.ScatterChart()
        chart_line.title = "Concentration vs Time"
        chart_line.style = 13
        chart_line.y_axis.title = "Conc_Total_Output (g/L)"
        chart_line.x_axis.title = "Time (min)"
        chart_line.x_axis.scaling.min = 0
        chart_line.height = 14
        chart_line.width = 22
        
        # Enable axis tick labels (benchmark numbers)
        chart_line.x_axis.tickLblPos = "low"      # Show X-axis numbers
        chart_line.y_axis.tickLblPos = "low"      # Show Y-axis numbers
        chart_line.x_axis.delete = False          # Ensure axis is visible
        chart_line.y_axis.delete = False
        chart_line.y_axis.numFmt = "0.00"         # Format: 2 decimal places
        chart_line.x_axis.numFmt = "0"            # Format: whole numbers
        
        # Colors for scenarios
        colors = ["0070C0", "C00000", "00B050", "7030A0", "FF6600", "00B0F0"]
        
        scenarios = final_df['Scenario_ID'].unique()
        current_row = 2
        bar_data = []  # Collect for bar chart
        
        for idx, sc_id in enumerate(scenarios):
            sc_df = final_df[final_df['Scenario_ID'] == sc_id]
            count = len(sc_df)
            
            data_x = openpyxl.chart.Reference(ws_data, min_col=2, min_row=current_row, max_row=current_row + count - 1)
            data_y = openpyxl.chart.Reference(ws_data, min_col=6, min_row=current_row, max_row=current_row + count - 1)
            
            series = openpyxl.chart.Series(data_y, xvalues=data_x, title=str(sc_id))
            series.marker.symbol = "none"
            series.smooth = True
            series.graphicalProperties.line.solidFill = colors[idx % len(colors)]
            series.graphicalProperties.line.width = 25000
            
            chart_line.series.append(series)
            bar_data.append((str(sc_id), float(sc_df['Conc_Total_Output_gL'].iloc[-1])))
            current_row += count
        
        # Add data labels to line chart (show Y values)
        from openpyxl.chart.label import DataLabelList
        
        ws_dash.add_chart(chart_line, "B3")
        
        # Write bar data to _ChartData sheet (columns H and I)
        ws_data.cell(row=1, column=8, value="Scenario")
        ws_data.cell(row=1, column=9, value="Final_Conc")
        for i, (sc_id, conc) in enumerate(bar_data):
            ws_data.cell(row=i+2, column=8, value=sc_id)
            ws_data.cell(row=i+2, column=9, value=conc)
        
        # Bar Chart
        chart_bar = openpyxl.chart.BarChart()
        chart_bar.title = "Final Concentration Comparison"
        chart_bar.style = 10
        chart_bar.y_axis.title = "Conc (g/L)"
        chart_bar.x_axis.title = "Scenario"
        chart_bar.height = 10
        chart_bar.width = 14
        
        # Enable axis tick labels (benchmark numbers)
        chart_bar.x_axis.tickLblPos = "low"
        chart_bar.y_axis.tickLblPos = "low"
        chart_bar.x_axis.delete = False
        chart_bar.y_axis.delete = False
        chart_bar.y_axis.numFmt = "0.00"
        
        # Reference data from _ChartData sheet (H and I columns)
        bar_categories = openpyxl.chart.Reference(ws_data, min_col=8, min_row=2, max_row=len(bar_data)+1)
        bar_values = openpyxl.chart.Reference(ws_data, min_col=9, min_row=1, max_row=len(bar_data)+1)
        
        chart_bar.add_data(bar_values, titles_from_data=True)
        chart_bar.set_categories(bar_categories)
        chart_bar.shape = 4
        
        # Add data labels to bar chart (show values on bars)
        chart_bar.dataLabels = DataLabelList()
        chart_bar.dataLabels.showVal = True
        chart_bar.dataLabels.showCatName = False
        
        ws_dash.add_chart(chart_bar, "B35")
        
        # Set Dashboard as active
        wb.active = wb.sheetnames.index("Dashboard")
        
        wb.save(OUTPUT_FILE)
        print("Success! Dashboard created.")
        
    except PermissionError:
        print(f"CRITICAL ERROR: {OUTPUT_FILE} is locked. Please close it.")
    except Exception as e:
        print(f"Error writing output/chart: {e}")

if __name__ == "__main__":
    run_simulation()
